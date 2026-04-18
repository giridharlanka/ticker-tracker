"""Portfolio XLSX report (Holdings, Summary, Metadata)."""

from __future__ import annotations

from collections.abc import Mapping
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ticker_tracker.calculator import purchase_amount_lines_by_ccy
from ticker_tracker.fx.base import FXRate

_HEADER_FONT = Font(bold=True)
_POS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_NEG_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_PCT_POS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_PCT_NEG_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _sheet_filename_timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def default_workbook_filename() -> str:
    return f"Ticker_summary_{_sheet_filename_timestamp()}.xlsx"


def _pct_fill(value: Any) -> PatternFill | None:
    if not isinstance(value, int | float):
        return None
    if value > 0:
        return _PCT_POS_FILL
    if value < 0:
        return _PCT_NEG_FILL
    return None


def _mixed_purchase_currencies(holdings_rows: list[dict[str, Any]]) -> bool:
    """True if rows use more than one *report_ccy* (totals of purchase-currency amounts would mix units)."""
    seen: set[str] = set()
    for h in holdings_rows:
        raw = h.get("report_ccy")
        if raw is None:
            continue
        s = str(raw).strip().upper()
        if not s or s == "—":
            continue
        seen.add(s)
    return len(seen) > 1


def build_portfolio_workbook(
    path: str | Path,
    *,
    base_currency: str,
    holdings_rows: list[dict[str, Any]],
    summary: Mapping[str, Any],
    metadata: Mapping[str, Any],
) -> Path:
    """
    Write a three-sheet workbook to *path*.

    *holdings_rows* must supply keys aligned with the engine (see ``engine.py``).
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_h = wb.active
    assert ws_h is not None
    ws_h.title = "Holdings"
    b = base_currency.upper()

    h1 = [
        "Ticker",
        "Shares",
        "Cost/sh (purch.)",
        "CCY",
        "Cost basis (purch.)",
        "Price/sh (purch.)",
        "Latest value (purch.)",
        "G/L (purch.)",
        "G/L %",
    ]
    for col, title in enumerate(h1, start=1):
        c = ws_h.cell(row=1, column=col, value=title)
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    pct_col = 9
    gl_amt_col = 8
    last_data = len(holdings_rows) + 1
    for r, holding in enumerate(holdings_rows, start=2):
        ws_h.cell(r, 1, holding.get("ticker"))
        ws_h.cell(r, 2, holding.get("shares"))
        ws_h.cell(r, 3, holding.get("cost_per_share_purchase"))
        ws_h.cell(r, 4, holding.get("report_ccy"))
        ws_h.cell(r, 5, holding.get("cost_basis_purchase"))
        ws_h.cell(r, 6, holding.get("price_per_share_purchase"))
        ws_h.cell(r, 7, holding.get("current_value_purchase"))
        gl_p = holding.get("gain_loss_purchase")
        ws_h.cell(r, gl_amt_col, gl_p)
        pct_val = holding.get("gain_loss_pct_purchase")
        ws_h.cell(r, pct_col, pct_val)
        if isinstance(gl_p, int | float):
            if gl_p > 0:
                ws_h.cell(r, gl_amt_col).fill = _POS_FILL
            elif gl_p < 0:
                ws_h.cell(r, gl_amt_col).fill = _NEG_FILL
        pf = _pct_fill(pct_val)
        if pf:
            ws_h.cell(r, pct_col).fill = pf

    if not _mixed_purchase_currencies(holdings_rows):
        total_row = last_data + 1
        ws_h.cell(total_row, 1, "TOTAL").font = _HEADER_FONT
        if last_data >= 2:
            ws_h.cell(total_row, 2, f"=SUM(B2:B{last_data})").font = _HEADER_FONT
            ws_h.cell(total_row, 5, f"=SUM(E2:E{last_data})").font = _HEADER_FONT
            ws_h.cell(total_row, 7, f"=SUM(G2:G{last_data})").font = _HEADER_FONT
            ws_h.cell(total_row, 8, f"=SUM(H2:H{last_data})").font = _HEADER_FONT

    for col in range(1, len(h1) + 1):
        ws_h.column_dimensions[get_column_letter(col)].width = 16

    ws_s = wb.create_sheet("Summary")
    ws_s["A1"] = f"Base currency: {b}"
    ws_s["A1"].font = Font(bold=True, size=14)
    ws_s["A3"] = "Metric"
    ws_s["B3"] = "Base"
    ws_s["C3"] = "Purchased"
    for c in (1, 2, 3):
        ws_s.cell(3, c).font = _HEADER_FONT

    r = 4
    bu = b.upper()

    def fmt_base_cell(v: Any) -> str | float:
        if isinstance(v, bool) or not isinstance(v, int | float):
            return v if v is not None else "—"
        return f"{float(v):,.2f} {bu}"

    cost_map: dict[str, float] = dict(summary.get("totals_purchase_cost_by_ccy") or {})
    val_map: dict[str, float] = dict(summary.get("totals_purchase_value_by_ccy") or {})
    gl_map: dict[str, float] = dict(summary.get("totals_purchase_gl_by_ccy") or {})

    if not cost_map and summary.get("totals_purchase_cost_formatted"):
        ws_s.cell(r, 1, "Total invested")
        ws_s.cell(r, 2, fmt_base_cell(summary.get("total_cost_basis_base")))
        ws_s.cell(r, 3, summary.get("totals_purchase_cost_formatted") or "—")
        r += 1
        ws_s.cell(r, 1, "Current value")
        ws_s.cell(r, 2, fmt_base_cell(summary.get("total_current_value_base")))
        ws_s.cell(r, 3, summary.get("totals_purchase_value_formatted") or "—")
        r += 1
        ws_s.cell(r, 1, "Gain / loss")
        ws_s.cell(r, 2, fmt_base_cell(summary.get("total_gain_loss_base")))
        ws_s.cell(r, 3, summary.get("totals_purchase_gl_formatted") or "—")
        r += 1
    else:
        for label, base_key, pmap in (
            ("Total invested", "total_cost_basis_base", cost_map),
            ("Current value", "total_current_value_base", val_map),
            ("Gain / loss", "total_gain_loss_base", gl_map),
        ):
            lines = purchase_amount_lines_by_ccy(pmap)
            ws_s.cell(r, 1, label)
            ws_s.cell(r, 2, fmt_base_cell(summary.get(base_key)))
            ws_s.cell(r, 3, lines[0])
            r += 1
            for ln in lines[1:]:
                ws_s.cell(r, 1, "")
                ws_s.cell(r, 2, "")
                ws_s.cell(r, 3, ln)
                r += 1

    tr = summary.get("total_return_pct")
    ws_s.cell(r, 1, "Total return %")
    ws_s.cell(r, 2, tr)
    ws_s.cell(r, 3, tr)
    r += 1

    ws_s.cell(r, 1, "Holdings count (distinct tickers)")
    ws_s.cell(r, 2, summary.get("distinct_ticker_count", summary.get("holding_count")))
    ws_s.cell(r, 3, "—")
    r += 1

    row = r + 1
    for note in summary.get("assumption_notes") or []:
        ws_s.cell(row, 1, note)
        ws_s.cell(row, 1).alignment = Alignment(wrap_text=True)
        row += 1

    ws_m = wb.create_sheet("Metadata")
    ws_m["A1"] = "Run timestamp (UTC)"
    ws_m["B1"] = metadata.get("run_timestamp_utc", "")
    ws_m["A2"] = "Base currency"
    ws_m["B2"] = b
    ws_m["A3"] = "FX source"
    ws_m["B3"] = metadata.get("fx_source", "")
    r = 5
    ws_m.cell(r, 1, "FX rates used").font = _HEADER_FONT
    r += 1
    for c, title in enumerate(["From", "To", "Rate", "Fetched At (UTC)", "Source"], start=1):
        ws_m.cell(r, c, title).font = _HEADER_FONT
    r += 1
    for fx in metadata.get("fx_rates") or []:
        if isinstance(fx, FXRate):
            ws_m.cell(r, 1, fx.from_currency)
            ws_m.cell(r, 2, fx.to_currency)
            ws_m.cell(r, 3, fx.rate)
            ws_m.cell(r, 4, fx.fetched_at.astimezone(UTC).isoformat())
            ws_m.cell(r, 5, fx.source)
        elif isinstance(fx, dict):
            ws_m.cell(r, 1, fx.get("from_currency"))
            ws_m.cell(r, 2, fx.get("to_currency"))
            ws_m.cell(r, 3, fx.get("rate"))
            ws_m.cell(r, 4, fx.get("fetched_at"))
            ws_m.cell(r, 5, fx.get("source"))
        r += 1

    r += 1
    ws_m.cell(r, 1, "Finance source per ticker").font = _HEADER_FONT
    r += 1
    ws_m.cell(r, 1, "Ticker").font = _HEADER_FONT
    ws_m.cell(r, 2, "Source").font = _HEADER_FONT
    r += 1
    for t, src in sorted((metadata.get("finance_source_by_ticker") or {}).items()):
        ws_m.cell(r, 1, t)
        ws_m.cell(r, 2, src)
        r += 1

    r += 1
    ws_m.cell(r, 1, "Price fetch failed (tickers)").font = _HEADER_FONT
    r += 1
    ws_m.cell(r, 1, ", ".join(metadata.get("price_fetch_failed") or []) or "—")
    r += 2
    ws_m.cell(r, 1, "FX rate unavailable (tickers)").font = _HEADER_FONT
    r += 1
    ws_m.cell(r, 1, ", ".join(metadata.get("fx_unavailable_tickers") or []) or "—")

    wb.save(path)
    return path

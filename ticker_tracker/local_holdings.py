"""Read holdings from local CSV/XLSX files."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

_REQUIRED_FIELDS = ("ticker", "shares", "cost_basis")
_ROW_FIELD_ORDER = (
    "ticker",
    "exchange",
    "shares",
    "cost_basis",
    "purchase_currency",
    "currency_override",
)


def _resolve_path(path: str) -> Path:
    candidate = Path(path).expanduser()
    if candidate.is_absolute():
        return candidate
    return (Path.cwd() / candidate).resolve()


def _row_fields(column_map: dict[str, str]) -> list[str]:
    for field in _REQUIRED_FIELDS:
        if field not in column_map:
            raise KeyError(f"column_map must include '{field}'")
    return [f for f in _ROW_FIELD_ORDER if f in column_map]


def _normalized_column_map_for_csv(column_map: dict[str, str]) -> dict[str, str]:
    return {k: str(v).strip().lower() for k, v in column_map.items()}


def _normalized_entry(entry: dict[str, Any], fields: list[str]) -> dict[str, Any]:
    row = {field: str(entry.get(field, "") or "").strip() for field in fields}
    if any(row.get(field, "").strip() for field in _REQUIRED_FIELDS):
        return row
    return {}


def _read_csv(path: Path, column_map: dict[str, str]) -> list[dict[str, Any]]:
    fields = _row_fields(column_map)
    header_map = _normalized_column_map_for_csv(column_map)
    out: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        if reader.fieldnames is None:
            return []
        normalized_src = {str(name).strip().lower(): name for name in reader.fieldnames}
        for src_header in header_map.values():
            if src_header not in normalized_src:
                raise KeyError(f"CSV header {src_header!r} not found in file.")
        for src_row in reader:
            row_raw = {
                field: src_row.get(normalized_src[header_map[field]], "") for field in fields
            }
            row = _normalized_entry(row_raw, fields)
            if row:
                out.append(row)
    return out


def _read_xlsx(path: Path, sheet_name: str, column_map: dict[str, str]) -> list[dict[str, Any]]:
    from ticker_tracker.google.sheets import column_letter_to_index

    fields = _row_fields(column_map)
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet tab {sheet_name!r} not found in {path.name}.")
    ws = wb[sheet_name]
    letters = [column_map[f].strip().upper() for f in fields]
    col_indices = [column_letter_to_index(letter) + 1 for letter in letters]
    lo, hi = min(col_indices), max(col_indices)
    out: list[dict[str, Any]] = []
    try:
        # read_only sheets often have max_row None; iter_rows streams all data rows.
        for row in ws.iter_rows(min_row=2, min_col=lo, max_col=hi, values_only=True):
            row_raw: dict[str, Any] = {}
            for field, col_idx in zip(fields, col_indices, strict=True):
                off = col_idx - lo
                row_raw[field] = row[off] if off < len(row) else None
            row_norm = _normalized_entry(row_raw, fields)
            if row_norm:
                out.append(row_norm)
    finally:
        wb.close()
    return out


def read_local_holdings(
    path: str,
    *,
    column_map: dict[str, str],
    sheet_name: str = "Holdings",
) -> list[dict[str, Any]]:
    resolved = _resolve_path(path)
    if not resolved.is_file():
        raise FileNotFoundError(str(resolved))
    ext = resolved.suffix.lower()
    if ext == ".csv":
        return _read_csv(resolved, column_map)
    if ext in {".xlsx", ".xlsm"}:
        return _read_xlsx(resolved, sheet_name, column_map)
    raise ValueError(f"Unsupported local holdings file type: {resolved.name}")

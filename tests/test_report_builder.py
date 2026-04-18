"""Tests for portfolio XLSX report builder."""

from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook
from ticker_tracker.fx.base import FXRate
from ticker_tracker.report_builder import build_portfolio_workbook


def test_build_portfolio_workbook_schema(tmp_path: Path) -> None:
    base = "SGD"
    holdings = [
        {
            "ticker": "AAPL",
            "shares": 2.0,
            "report_ccy": "SGD",
            "cost_per_share_purchase": 250.0,
            "cost_basis_purchase": 500.0,
            "price_per_share_purchase": 135.0,
            "current_value_purchase": 270.0,
            "gain_loss_purchase": -230.0,
            "gain_loss_pct_purchase": -46.0,
            "cost_basis_base": 500.0,
            "native_ccy": "USD",
            "price_native": 100.0,
            "fx_rate_display": 1.35,
            "current_price_base": 135.0,
            "current_value_base": 270.0,
            "gain_loss_base": -230.0,
            "gain_loss_pct": -46.0,
        }
    ]
    summary = {
        "total_cost_basis_base": 500.0,
        "total_current_value_base": 270.0,
        "total_gain_loss_base": -230.0,
        "total_return_pct": -46.0,
        "holding_count": 1,
        "distinct_ticker_count": 1,
        "totals_purchase_cost_formatted": "SGD 500.00",
        "totals_purchase_value_formatted": "SGD 270.00",
        "totals_purchase_gl_formatted": "SGD -230.00",
        "totals_purchase_cost_by_ccy": {"SGD": 500.0},
        "totals_purchase_value_by_ccy": {"SGD": 270.0},
        "totals_purchase_gl_by_ccy": {"SGD": -230.0},
        "markets_covered": ["US"],
        "currencies_in_portfolio": ["USD"],
        "assumption_notes": ["Note A"],
    }
    fx = FXRate("SGD", "USD", 0.74, datetime(2026, 1, 1, tzinfo=UTC), "frankfurter")
    metadata = {
        "run_timestamp_utc": "2026-01-01T00:00:00+00:00",
        "fx_source": "frankfurter",
        "fx_rates": [fx],
        "finance_source_by_ticker": {"AAPL": "yahoo"},
        "price_fetch_failed": [],
        "fx_unavailable_tickers": [],
    }
    out = tmp_path / "Ticker_summary_test.xlsx"
    build_portfolio_workbook(
        out,
        base_currency=base,
        holdings_rows=holdings,
        summary=summary,
        metadata=metadata,
    )
    wb = load_workbook(out)
    assert "Holdings" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    assert "Metadata" in wb.sheetnames
    ws = wb["Holdings"]
    assert "CCY" in str(ws.cell(1, 4).value)
    assert ws.cell(2, 1).value == "AAPL"
    assert ws.cell(2, 4).value == "SGD"
    assert ws.cell(2, 9).value == -46.0
    assert getattr(ws.cell(2, 9).fill, "patternType", None) == "solid"
    assert ws.cell(3, 1).value == "TOTAL"
    assert ws.cell(3, 5).value == "=SUM(E2:E2)"
    assert ws.cell(3, 7).value == "=SUM(G2:G2)"
    ws_sum = wb["Summary"]
    assert ws_sum.cell(3, 2).value == "Base"
    assert "SGD" in str(ws_sum.cell(4, 2).value)


def test_summary_purchased_splits_multiple_currencies(tmp_path: Path) -> None:
    holdings = [
        {
            "ticker": "X",
            "shares": 1.0,
            "report_ccy": "USD",
            "cost_per_share_purchase": 1.0,
            "cost_basis_purchase": 1.0,
            "price_per_share_purchase": 1.0,
            "current_value_purchase": 1.0,
            "gain_loss_purchase": 0.0,
            "gain_loss_pct_purchase": 0.0,
            "cost_basis_base": 1.0,
            "native_ccy": "USD",
            "price_native": 1.0,
            "fx_rate_display": 1.0,
            "current_price_base": 1.0,
            "current_value_base": 1.0,
            "gain_loss_base": 0.0,
            "gain_loss_pct": 0.0,
        }
    ]
    summary = {
        "total_cost_basis_base": 300.0,
        "total_current_value_base": 400.0,
        "total_gain_loss_base": 100.0,
        "total_return_pct": 10.0,
        "distinct_ticker_count": 1,
        "totals_purchase_cost_by_ccy": {"HKD": 100.0, "USD": 200.0},
        "totals_purchase_value_by_ccy": {"HKD": 120.0, "USD": 280.0},
        "totals_purchase_gl_by_ccy": {"HKD": 20.0, "USD": 80.0},
        "assumption_notes": [],
    }
    out = tmp_path / "multi.xlsx"
    build_portfolio_workbook(
        out,
        base_currency="SGD",
        holdings_rows=holdings,
        summary=summary,
        metadata={
            "fx_rates": [],
            "finance_source_by_ticker": {},
            "price_fetch_failed": [],
            "fx_unavailable_tickers": [],
        },
    )
    ws = load_workbook(out)["Summary"]
    assert ws.cell(4, 1).value == "Total invested"
    assert ws.cell(5, 1).value in (None, "")
    assert "HKD" in str(ws.cell(4, 3).value)
    assert "USD" in str(ws.cell(5, 3).value)


def test_holdings_sheet_omits_total_row_when_mixed_purchase_currencies(tmp_path: Path) -> None:
    """TOTAL sums purchase-currency columns; omit when CCY column mixes units."""
    row = {
        "ticker": "A",
        "shares": 1.0,
        "report_ccy": "USD",
        "cost_per_share_purchase": 1.0,
        "cost_basis_purchase": 1.0,
        "price_per_share_purchase": 1.0,
        "current_value_purchase": 1.0,
        "gain_loss_purchase": 0.0,
        "gain_loss_pct_purchase": 0.0,
        "cost_basis_base": 1.0,
        "native_ccy": "USD",
        "price_native": 1.0,
        "fx_rate_display": 1.0,
        "current_price_base": 1.0,
        "current_value_base": 1.0,
        "gain_loss_base": 0.0,
        "gain_loss_pct": 0.0,
    }
    row_b = {**row, "ticker": "B", "report_ccy": "SGD"}
    holdings = [row, row_b]
    summary = {
        "total_cost_basis_base": 2.0,
        "total_current_value_base": 2.0,
        "total_gain_loss_base": 0.0,
        "total_return_pct": 0.0,
        "distinct_ticker_count": 2,
        "totals_purchase_cost_by_ccy": {"USD": 1.0, "SGD": 1.0},
        "totals_purchase_value_by_ccy": {"USD": 1.0, "SGD": 1.0},
        "totals_purchase_gl_by_ccy": {"USD": 0.0, "SGD": 0.0},
        "assumption_notes": [],
    }
    out = tmp_path / "mixed_ccy_holdings.xlsx"
    build_portfolio_workbook(
        out,
        base_currency="USD",
        holdings_rows=holdings,
        summary=summary,
        metadata={
            "fx_rates": [],
            "finance_source_by_ticker": {},
            "price_fetch_failed": [],
            "fx_unavailable_tickers": [],
        },
    )
    ws = load_workbook(out)["Holdings"]
    assert ws.cell(2, 1).value == "A"
    assert ws.cell(3, 1).value == "B"
    assert ws.cell(4, 1).value is None
    for col in range(1, 10):
        assert ws.cell(4, col).value is None
    assert all(ws.cell(3, c).value != "TOTAL" for c in range(1, 10))

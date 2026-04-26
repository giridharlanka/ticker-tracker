"""Tests for local CSV/XLSX holdings readers."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from ticker_tracker.local_holdings import read_local_holdings


def test_read_local_csv_by_headers(tmp_path: Path) -> None:
    csv_path = tmp_path / "holdings.csv"
    csv_path.write_text(
        "symbol,qty,cost,exchange\nAAPL,10,100,NYSE\nMSFT,5,200,NASDAQ\n",
        encoding="utf-8",
    )
    rows = read_local_holdings(
        str(csv_path),
        column_map={
            "ticker": "symbol",
            "shares": "qty",
            "cost_basis": "cost",
            "exchange": "exchange",
        },
    )
    assert len(rows) == 2
    assert rows[0]["ticker"] == "AAPL"
    assert rows[1]["shares"] == "5"


def test_read_local_xlsx_by_column_letters(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "holdings.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Holdings"
    ws["A1"] = "ticker"
    ws["B1"] = "shares"
    ws["C1"] = "cost_basis"
    ws["A2"] = "VOO"
    ws["B2"] = 2
    ws["C2"] = 500
    wb.save(xlsx_path)
    rows = read_local_holdings(
        str(xlsx_path),
        column_map={"ticker": "A", "shares": "B", "cost_basis": "C"},
        sheet_name="Holdings",
    )
    assert rows == [{"ticker": "VOO", "shares": "2", "cost_basis": "500"}]
